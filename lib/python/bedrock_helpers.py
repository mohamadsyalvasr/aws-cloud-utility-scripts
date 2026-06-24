"""
bedrock_helpers.py
Shared utility functions for Bedrock observability report generation.
"""

import subprocess
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# =============================================================================
# Constants (shared across modules)
# =============================================================================
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11, color="333333")
CONCLUSION_FONT = Font(italic=True, size=10, color="4472C4")
DATA_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


# =============================================================================
# Data Helpers
# =============================================================================
def datapoints_to_df(datapoints, value_col='Sum'):
    """Convert CloudWatch datapoints list to a sorted DataFrame."""
    if not datapoints:
        return pd.DataFrame()
    df = pd.DataFrame(datapoints)
    if 'Timestamp' not in df.columns:
        return pd.DataFrame()
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])
    df = df.sort_values('Timestamp').reset_index(drop=True)
    return df


def safe_sum(datapoints, col='Sum'):
    """Safely sum a metric's datapoints."""
    if not datapoints:
        return 0
    total = 0
    for dp in datapoints:
        val = dp.get(col)
        if val is not None:
            total += val
    return total


def safe_avg(datapoints, col='Average'):
    """Safely average a metric's datapoints."""
    if not datapoints:
        return 0
    vals = [dp.get(col, 0) or 0 for dp in datapoints if dp.get(col) is not None]
    return sum(vals) / len(vals) if vals else 0


def short_model_name(model_id):
    """Shorten model ID for display."""
    if '/' in model_id:
        return model_id.split('/')[-1]
    if ':' in model_id:
        return model_id.split(':')[-1]
    return model_id[:40]


def get_aws_account_info():
    """Retrieve Account ID and Account Name from AWS CLI."""
    account_id = "UnknownAccount"
    account_name = ""
    try:
        r = subprocess.run(["aws", "sts", "get-caller-identity", "--query", "Account", "--output", "text"],
                           capture_output=True, text=True, timeout=30)
        if r.returncode == 0 and r.stdout.strip():
            account_id = r.stdout.strip()
    except Exception:
        pass
    try:
        r = subprocess.run(["aws", "iam", "list-account-aliases", "--query", "AccountAliases[0]", "--output", "text"],
                           capture_output=True, text=True, timeout=30)
        if r.returncode == 0 and r.stdout.strip() not in ("", "None"):
            account_name = r.stdout.strip()
    except Exception:
        pass
    return account_id, account_name


# =============================================================================
# Excel Helpers
# =============================================================================
def write_section_header(ws, row, title):
    """Write a section title row."""
    ws.cell(row=row, column=1, value=title).font = SUBTITLE_FONT
    return row + 1


def write_table(ws, start_row, headers, data_rows):
    """Write a formatted table with headers and data."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = DATA_BORDER
    for r_idx, row_data in enumerate(data_rows, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = DATA_BORDER
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row_data in data_rows:
            if col - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col - 1])))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 45)
    return start_row + len(data_rows) + 1


def insert_chart_image(ws, row, col, chart_bytes, width_px=850):
    """Insert a chart image into the worksheet."""
    if chart_bytes is None:
        return row
    img = XLImage(chart_bytes)
    img.width = width_px
    img.height = int(width_px * 0.45)
    ws.add_image(img, f'{get_column_letter(col)}{row}')
    rows_needed = int(img.height / 15) + 2
    return row + rows_needed


def write_conclusion(ws, row, conclusions):
    """Write conclusion lines at the bottom of a sheet."""
    row += 1
    ws.cell(row=row, column=1, value="── Conclusion ──").font = SUBTITLE_FONT
    row += 1
    for line in conclusions:
        ws.cell(row=row, column=1, value=f"• {line}").font = CONCLUSION_FONT
        row += 1
    return row
