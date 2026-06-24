"""
generate_usage_charts.py
Generates a comprehensive Excel report with embedded charts for Bedrock
observability metrics — matching CloudWatch GenAI Observability dashboard.

Sheets:
  1. Token Usage          - Input/Output tokens, Total tokens, Prompt Cache
  2. Latency & Performance - InvocationLatency, TTFT, EstimatedTPMQuotaUsage
  3. Volume & Distribution - Invocation count, Input token size distribution
  4. Reliability & Errors  - Throttles, Client vs Server errors

Usage:
    python3 generate_usage_charts.py <metrics_json> <output_dir> <start_date> <end_date>

Output:
    <output_dir>/Bedrock_Observability_Report.xlsx
    <output_dir>/bedrock_usage_report.zip
"""

import json
import os
import sys
import zipfile
import subprocess
from datetime import datetime
from io import BytesIO

import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# Constants
# =============================================================================
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11, color="333333")
DATA_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
COLORS = ['#4472C4', '#ED7D31', '#A5A5A5', '#FFC000', '#5B9BD5',
           '#70AD47', '#264478', '#9B59B6', '#E74C3C', '#1ABC9C']


# =============================================================================
# AWS Account Info
# =============================================================================
def get_aws_account_info():
    """Retrieve Account ID and Account Name from AWS CLI."""
    account_id = "UnknownAccount"
    account_name = ""
    try:
        r = subprocess.run(["aws", "sts", "get-caller-identity", "--query", "Account", "--output", "text"],
                           capture_output=True, text=True, timeout=30)
        if r.returncode == 0 and r.stdout.strip():
            account_id = r.stdout.strip()
    except Exception:
        pass
    try:
        r = subprocess.run(["aws", "iam", "list-account-aliases", "--query", "AccountAliases[0]", "--output", "text"],
                           capture_output=True, text=True, timeout=30)
        if r.returncode == 0 and r.stdout.strip() not in ("", "None"):
            account_name = r.stdout.strip()
    except Exception:
        pass
    return account_id, account_name


# =============================================================================
# Helper: Parse datapoints into a sorted DataFrame
# =============================================================================
def datapoints_to_df(datapoints, value_col='Sum'):
    """Convert CloudWatch datapoints list to a sorted DataFrame."""
    if not datapoints:
        return pd.DataFrame()
    df = pd.DataFrame(datapoints)
    if 'Timestamp' not in df.columns:
        return pd.DataFrame()
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])
    df = df.sort_values('Timestamp').reset_index(drop=True)
    return df


def safe_sum(datapoints, col='Sum'):
    """Safely sum a metric's datapoints."""
    if not datapoints:
        return 0
    return sum(dp.get(col, 0) or 0 for dp in datapoints)


def safe_avg(datapoints, col='Average'):
    """Safely average a metric's datapoints."""
    if not datapoints:
        return 0
    vals = [dp.get(col, 0) or 0 for dp in datapoints if dp.get(col) is not None]
    return sum(vals) / len(vals) if vals else 0


def short_model_name(model_id):
    """Shorten model ID for display."""
    if '/' in model_id:
        return model_id.split('/')[-1]
    if ':' in model_id:
        return model_id.split(':')[-1]
    return model_id[:40]


# =============================================================================
# Chart Generation Functions
# =============================================================================
def chart_to_bytes(fig, dpi=130):
    """Convert matplotlib figure to PNG bytes."""
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


def make_token_usage_chart(models_data):
    """Bar chart: Input vs Output tokens per model."""
    labels = [short_model_name(m['model_id']) for m in models_data]
    input_vals = [safe_sum(m['metrics'].get('input_tokens', [])) for m in models_data]
    output_vals = [safe_sum(m['metrics'].get('output_tokens', [])) for m in models_data]

    fig, ax = plt.subplots(figsize=(12, max(4, len(labels) * 0.6)))
    y = range(len(labels))
    h = 0.35
    ax.barh([i - h/2 for i in y], input_vals, h, label='Input Tokens', color=COLORS[0])
    ax.barh([i + h/2 for i in y], output_vals, h, label='Output Tokens', color=COLORS[1])
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel('Token Count')
    ax.set_title('Input and Output Token Usage by Model', fontweight='bold')
    ax.legend(loc='lower right')
    ax.grid(axis='x', alpha=0.3)
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_token_trend_chart(models_data):
    """Line chart: Total tokens over time per model."""
    fig, ax = plt.subplots(figsize=(12, 5))
    for idx, m in enumerate(models_data):
        label = short_model_name(m['model_id'])
        input_dp = m['metrics'].get('input_tokens', [])
        output_dp = m['metrics'].get('output_tokens', [])
        if not input_dp:
            continue
        df_in = datapoints_to_df(input_dp)
        df_out = datapoints_to_df(output_dp)
        if df_in.empty:
            continue
        # Merge and sum
        df_in = df_in.rename(columns={'Sum': 'InputSum'})
        if not df_out.empty:
            df_out = df_out.rename(columns={'Sum': 'OutputSum'})
            merged = pd.merge(df_in[['Timestamp', 'InputSum']], df_out[['Timestamp', 'OutputSum']],
                              on='Timestamp', how='outer').fillna(0)
            merged['Total'] = merged['InputSum'] + merged['OutputSum']
        else:
            merged = df_in[['Timestamp', 'InputSum']].copy()
            merged['Total'] = merged['InputSum']
        merged = merged.sort_values('Timestamp')
        ax.plot(merged['Timestamp'], merged['Total'], marker='o', markersize=3,
                label=label, color=COLORS[idx % len(COLORS)], linewidth=1.5)
    ax.set_title('Total Tokens Over Time', fontweight='bold')
    ax.set_xlabel('Date')
    ax.set_ylabel('Total Tokens')
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
    plt.xticks(rotation=45)
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_cache_chart(models_data):
    """Bar chart: Cache Read vs Write tokens."""
    labels, reads, writes = [], [], []
    for m in models_data:
        r = safe_sum(m['metrics'].get('cache_read_tokens', []))
        w = safe_sum(m['metrics'].get('cache_write_tokens', []))
        if r > 0 or w > 0:
            labels.append(short_model_name(m['model_id']))
            reads.append(r)
            writes.append(w)
    if not labels:
        return None
    fig, ax = plt.subplots(figsize=(10, max(3, len(labels) * 0.5)))
    y = range(len(labels))
    h = 0.35
    ax.barh([i - h/2 for i in y], reads, h, label='Cache Read', color=COLORS[4])
    ax.barh([i + h/2 for i in y], writes, h, label='Cache Write', color=COLORS[5])
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel('Token Count')
    ax.set_title('Prompt Cache Usage (Read vs Write Tokens)', fontweight='bold')
    ax.legend()
    ax.grid(axis='x', alpha=0.3)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_latency_chart(models_data):
    """Line chart: Average InvocationLatency over time."""
    fig, ax = plt.subplots(figsize=(12, 5))
    for idx, m in enumerate(models_data):
        dp = m['metrics'].get('invocation_latency', [])
        if not dp:
            continue
        df = datapoints_to_df(dp, 'Average')
        if df.empty or 'Average' not in df.columns:
            continue
        label = short_model_name(m['model_id'])
        ax.plot(df['Timestamp'], df['Average'], marker='o', markersize=3,
                label=label, color=COLORS[idx % len(COLORS)], linewidth=1.5)
    ax.set_title('End-to-End Invocation Latency (Avg ms)', fontweight='bold')
    ax.set_xlabel('Date')
    ax.set_ylabel('Latency (ms)')
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_ttft_chart(models_data):
    """Line chart: Time To First Token over time."""
    fig, ax = plt.subplots(figsize=(12, 5))
    has_data = False
    for idx, m in enumerate(models_data):
        dp = m['metrics'].get('time_to_first_token', [])
        if not dp:
            continue
        df = datapoints_to_df(dp, 'Average')
        if df.empty or 'Average' not in df.columns:
            continue
        has_data = True
        label = short_model_name(m['model_id'])
        ax.plot(df['Timestamp'], df['Average'], marker='s', markersize=3,
                label=label, color=COLORS[idx % len(COLORS)], linewidth=1.5)
    if not has_data:
        plt.close(fig)
        return None
    ax.set_title('Time To First Token - TTFT (Avg ms)', fontweight='bold')
    ax.set_xlabel('Date')
    ax.set_ylabel('TTFT (ms)')
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_tpm_chart(models_data):
    """Line chart: Estimated TPM Quota Usage over time."""
    fig, ax = plt.subplots(figsize=(12, 5))
    has_data = False
    for idx, m in enumerate(models_data):
        dp = m['metrics'].get('estimated_tpm_quota', [])
        if not dp:
            continue
        df = datapoints_to_df(dp, 'Maximum')
        if df.empty or 'Maximum' not in df.columns:
            continue
        has_data = True
        label = short_model_name(m['model_id'])
        ax.plot(df['Timestamp'], df['Maximum'], marker='^', markersize=3,
                label=label, color=COLORS[idx % len(COLORS)], linewidth=1.5)
    if not has_data:
        plt.close(fig)
        return None
    ax.set_title('Estimated TPM Quota Usage (Max)', fontweight='bold')
    ax.set_xlabel('Date')
    ax.set_ylabel('Tokens Per Minute')
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_invocation_count_chart(models_data):
    """Line chart: Invocation count over time."""
    fig, ax = plt.subplots(figsize=(12, 5))
    for idx, m in enumerate(models_data):
        dp = m['metrics'].get('invocations', [])
        if not dp:
            continue
        df = datapoints_to_df(dp)
        if df.empty or 'Sum' not in df.columns:
            continue
        label = short_model_name(m['model_id'])
        ax.plot(df['Timestamp'], df['Sum'], marker='o', markersize=3,
                label=label, color=COLORS[idx % len(COLORS)], linewidth=1.5)
    ax.set_title('Invocation Count Over Time', fontweight='bold')
    ax.set_xlabel('Date')
    ax.set_ylabel('Invocations')
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_token_distribution_chart(models_data):
    """Box-like chart: Input token size distribution per model."""
    fig, ax = plt.subplots(figsize=(10, max(4, len(models_data) * 0.6)))
    labels, avgs, mins, maxs = [], [], [], []
    for m in models_data:
        dp = m['metrics'].get('input_tokens', [])
        if not dp:
            continue
        avg_vals = [d.get('Average', 0) or 0 for d in dp if d.get('Average') is not None]
        min_vals = [d.get('Minimum', 0) or 0 for d in dp if d.get('Minimum') is not None]
        max_vals = [d.get('Maximum', 0) or 0 for d in dp if d.get('Maximum') is not None]
        if not avg_vals:
            continue
        labels.append(short_model_name(m['model_id']))
        avgs.append(sum(avg_vals) / len(avg_vals))
        mins.append(min(min_vals) if min_vals else 0)
        maxs.append(max(max_vals) if max_vals else 0)
    if not labels:
        plt.close(fig)
        return None
    y = range(len(labels))
    # Plot as horizontal error bars (min → max) with dot at average
    for i in range(len(labels)):
        ax.plot([mins[i], maxs[i]], [i, i], color=COLORS[2], linewidth=2, solid_capstyle='round')
        ax.plot(avgs[i], i, 'o', color=COLORS[0], markersize=8)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel('Input Tokens per Request')
    ax.set_title('Request Distribution by Input Token Size (Min | Avg | Max)', fontweight='bold')
    ax.grid(axis='x', alpha=0.3)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_throttles_chart(models_data):
    """Line chart: Throttles over time."""
    fig, ax = plt.subplots(figsize=(12, 5))
    has_data = False
    for idx, m in enumerate(models_data):
        dp = m['metrics'].get('throttles', [])
        if not dp:
            continue
        df = datapoints_to_df(dp)
        if df.empty or 'Sum' not in df.columns:
            continue
        if df['Sum'].sum() == 0:
            continue
        has_data = True
        label = short_model_name(m['model_id'])
        ax.plot(df['Timestamp'], df['Sum'], marker='x', markersize=4,
                label=label, color=COLORS[idx % len(COLORS)], linewidth=1.5)
    if not has_data:
        plt.close(fig)
        return None
    ax.set_title('Invocation Throttles Over Time', fontweight='bold')
    ax.set_xlabel('Date')
    ax.set_ylabel('Throttle Count')
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_errors_chart(models_data):
    """Stacked bar: Client vs Server errors per model."""
    labels, client_errs, server_errs = [], [], []
    for m in models_data:
        ce = safe_sum(m['metrics'].get('client_errors', []))
        se = safe_sum(m['metrics'].get('server_errors', []))
        if ce > 0 or se > 0:
            labels.append(short_model_name(m['model_id']))
            client_errs.append(ce)
            server_errs.append(se)
    if not labels:
        return None
    fig, ax = plt.subplots(figsize=(10, max(4, len(labels) * 0.5)))
    y = range(len(labels))
    ax.barh(y, client_errs, label='Client Errors (4xx)', color=COLORS[1])
    ax.barh(y, server_errs, left=client_errs, label='Server Errors (5xx)', color=COLORS[8] if len(COLORS) > 8 else 'red')
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel('Error Count')
    ax.set_title('Invocation Errors (Client vs Server)', fontweight='bold')
    ax.legend()
    ax.grid(axis='x', alpha=0.3)
    plt.tight_layout()
    return chart_to_bytes(fig)


def make_errors_trend_chart(models_data):
    """Line chart: Client + Server errors over time (combined)."""
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    has_client, has_server = False, False
    for idx, m in enumerate(models_data):
        label = short_model_name(m['model_id'])
        color = COLORS[idx % len(COLORS)]
        # Client errors
        dp = m['metrics'].get('client_errors', [])
        if dp:
            df = datapoints_to_df(dp)
            if not df.empty and 'Sum' in df.columns and df['Sum'].sum() > 0:
                has_client = True
                ax1.plot(df['Timestamp'], df['Sum'], marker='x', markersize=3,
                         label=label, color=color, linewidth=1.5)
        # Server errors
        dp = m['metrics'].get('server_errors', [])
        if dp:
            df = datapoints_to_df(dp)
            if not df.empty and 'Sum' in df.columns and df['Sum'].sum() > 0:
                has_server = True
                ax2.plot(df['Timestamp'], df['Sum'], marker='x', markersize=3,
                         label=label, color=color, linewidth=1.5)
    if not has_client and not has_server:
        plt.close(fig)
        return None
    ax1.set_title('Client Errors (4xx) Over Time', fontweight='bold')
    ax1.set_ylabel('Count')
    ax1.legend(fontsize=8)
    ax1.grid(alpha=0.3)
    ax2.set_title('Server Errors (5xx) Over Time', fontweight='bold')
    ax2.set_xlabel('Date')
    ax2.set_ylabel('Count')
    ax2.legend(fontsize=8)
    ax2.grid(alpha=0.3)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    return chart_to_bytes(fig)


# =============================================================================
# Excel Workbook Generation
# =============================================================================
def write_header(ws, row, title):
    """Write a section title row."""
    ws.cell(row=row, column=1, value=title).font = SUBTITLE_FONT
    return row + 1


def write_table(ws, start_row, headers, data_rows):
    """Write a formatted table with headers and data."""
    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = DATA_BORDER
    # Data
    for r_idx, row_data in enumerate(data_rows, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = DATA_BORDER
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row_data in data_rows:
            if col - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col - 1])))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 45)
    return start_row + len(data_rows) + 1


def insert_chart_image(ws, row, col, chart_bytes, width_px=850):
    """Insert a chart image into the worksheet."""
    if chart_bytes is None:
        return row
    img = XLImage(chart_bytes)
    img.width = width_px
    img.height = int(width_px * 0.45)
    ws.add_image(img, f'{get_column_letter(col)}{row}')
    # Calculate rows needed (approx 15px per row)
    rows_needed = int(img.height / 15) + 2
    return row + rows_needed


def build_sheet_token_usage(wb, bedrock_models, account_label, period_str):
    """Sheet 1: Token Usage."""
    ws = wb.create_sheet("Token Usage")
    ws.cell(row=1, column=1, value=f"Bedrock Token Usage — {account_label}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period_str}").font = SUBTITLE_FONT
    row = 4

    # Summary table
    headers = ['Model', 'Region', 'Input Tokens', 'Output Tokens', 'Total Tokens',
               'Cache Read', 'Cache Write', 'Invocations']
    data = []
    for m in bedrock_models:
        inp = safe_sum(m['metrics'].get('input_tokens', []))
        out = safe_sum(m['metrics'].get('output_tokens', []))
        cr = safe_sum(m['metrics'].get('cache_read_tokens', []))
        cw = safe_sum(m['metrics'].get('cache_write_tokens', []))
        inv = safe_sum(m['metrics'].get('invocations', m['metrics'].get('model_invocations', [])))
        data.append([short_model_name(m['model_id']), m['region'],
                     int(inp), int(out), int(inp + out), int(cr), int(cw), int(inv)])
    row = write_table(ws, row, headers, data)
    row += 2

    # Charts
    row = write_header(ws, row, "Input & Output Token Usage")
    chart1 = make_token_usage_chart(bedrock_models)
    row = insert_chart_image(ws, row, 1, chart1)
    row += 1

    row = write_header(ws, row, "Total Tokens Over Time")
    chart2 = make_token_trend_chart(bedrock_models)
    row = insert_chart_image(ws, row, 1, chart2)
    row += 1

    row = write_header(ws, row, "Prompt Cache Usage (Read vs Write)")
    chart3 = make_cache_chart(bedrock_models)
    if chart3:
        row = insert_chart_image(ws, row, 1, chart3)
    else:
        ws.cell(row=row, column=1, value="No prompt cache data available.").font = SUBTITLE_FONT
        row += 2

    return ws


def build_sheet_latency(wb, bedrock_models, account_label, period_str):
    """Sheet 2: Latency & Performance."""
    ws = wb.create_sheet("Latency & Performance")
    ws.cell(row=1, column=1, value=f"Bedrock Latency & Performance — {account_label}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period_str}").font = SUBTITLE_FONT
    row = 4

    # Summary table
    headers = ['Model', 'Region', 'Avg Latency (ms)', 'Min Latency', 'Max Latency',
               'Avg TTFT (ms)', 'Max TPM Quota']
    data = []
    for m in bedrock_models:
        lat = m['metrics'].get('invocation_latency', m['metrics'].get('model_latency', []))
        ttft = m['metrics'].get('time_to_first_token', [])
        tpm = m['metrics'].get('estimated_tpm_quota', [])
        avg_lat = safe_avg(lat, 'Average')
        min_lat = min((d.get('Minimum', 0) or 0 for d in lat), default=0) if lat else 0
        max_lat = max((d.get('Maximum', 0) or 0 for d in lat), default=0) if lat else 0
        avg_ttft = safe_avg(ttft, 'Average')
        max_tpm = max((d.get('Maximum', 0) or 0 for d in tpm), default=0) if tpm else 0
        data.append([short_model_name(m['model_id']), m['region'],
                     round(avg_lat, 1), round(min_lat, 1), round(max_lat, 1),
                     round(avg_ttft, 1), int(max_tpm)])
    row = write_table(ws, row, headers, data)
    row += 2

    # Charts
    row = write_header(ws, row, "End-to-End Invocation Latency")
    chart1 = make_latency_chart(bedrock_models)
    row = insert_chart_image(ws, row, 1, chart1)
    row += 1

    row = write_header(ws, row, "Time To First Token (TTFT)")
    chart2 = make_ttft_chart(bedrock_models)
    if chart2:
        row = insert_chart_image(ws, row, 1, chart2)
    else:
        ws.cell(row=row, column=1, value="No TTFT data (streaming APIs only).").font = SUBTITLE_FONT
        row += 2
    row += 1

    row = write_header(ws, row, "Estimated TPM Quota Usage")
    chart3 = make_tpm_chart(bedrock_models)
    if chart3:
        row = insert_chart_image(ws, row, 1, chart3)
    else:
        ws.cell(row=row, column=1, value="No TPM quota data available.").font = SUBTITLE_FONT
        row += 2

    return ws


def build_sheet_volume(wb, bedrock_models, account_label, period_str):
    """Sheet 3: Volume & Distribution."""
    ws = wb.create_sheet("Volume & Distribution")
    ws.cell(row=1, column=1, value=f"Bedrock Volume & Distribution — {account_label}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period_str}").font = SUBTITLE_FONT
    row = 4

    # Summary table
    headers = ['Model', 'Region', 'Total Invocations', 'Avg Input Tokens/Req',
               'Min Input Tokens', 'Max Input Tokens']
    data = []
    for m in bedrock_models:
        inv = safe_sum(m['metrics'].get('invocations', m['metrics'].get('model_invocations', [])))
        inp_dp = m['metrics'].get('input_tokens', [])
        avg_inp = safe_avg(inp_dp, 'Average')
        min_inp = min((d.get('Minimum', 0) or 0 for d in inp_dp), default=0) if inp_dp else 0
        max_inp = max((d.get('Maximum', 0) or 0 for d in inp_dp), default=0) if inp_dp else 0
        data.append([short_model_name(m['model_id']), m['region'],
                     int(inv), round(avg_inp, 0), int(min_inp), int(max_inp)])
    row = write_table(ws, row, headers, data)
    row += 2

    # Charts
    row = write_header(ws, row, "Invocation Count Over Time")
    chart1 = make_invocation_count_chart(bedrock_models)
    row = insert_chart_image(ws, row, 1, chart1)
    row += 1

    row = write_header(ws, row, "Request Distribution by Input Token Size")
    chart2 = make_token_distribution_chart(bedrock_models)
    if chart2:
        row = insert_chart_image(ws, row, 1, chart2)
    else:
        ws.cell(row=row, column=1, value="No distribution data available.").font = SUBTITLE_FONT
        row += 2

    return ws


def build_sheet_reliability(wb, bedrock_models, account_label, period_str):
    """Sheet 4: Reliability & Errors."""
    ws = wb.create_sheet("Reliability & Errors")
    ws.cell(row=1, column=1, value=f"Bedrock Reliability & Errors — {account_label}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period_str}").font = SUBTITLE_FONT
    row = 4

    # Summary table
    headers = ['Model', 'Region', 'Invocations', 'Throttles', 'Client Errors',
               'Server Errors', 'Error Rate %']
    data = []
    for m in bedrock_models:
        inv = safe_sum(m['metrics'].get('invocations', m['metrics'].get('model_invocations', [])))
        thr = safe_sum(m['metrics'].get('throttles', []))
        ce = safe_sum(m['metrics'].get('client_errors', []))
        se = safe_sum(m['metrics'].get('server_errors', []))
        total_attempts = inv + thr + ce + se
        err_rate = ((ce + se) / total_attempts * 100) if total_attempts > 0 else 0
        data.append([short_model_name(m['model_id']), m['region'],
                     int(inv), int(thr), int(ce), int(se), round(err_rate, 2)])
    row = write_table(ws, row, headers, data)
    row += 2

    # Charts
    row = write_header(ws, row, "Invocation Throttles Over Time")
    chart1 = make_throttles_chart(bedrock_models)
    if chart1:
        row = insert_chart_image(ws, row, 1, chart1)
    else:
        ws.cell(row=row, column=1, value="No throttles in this period. ✅").font = SUBTITLE_FONT
        row += 2
    row += 1

    row = write_header(ws, row, "Invocation Errors (Client vs Server)")
    chart2 = make_errors_chart(bedrock_models)
    if chart2:
        row = insert_chart_image(ws, row, 1, chart2)
    else:
        ws.cell(row=row, column=1, value="No errors in this period. ✅").font = SUBTITLE_FONT
        row += 2
    row += 1

    row = write_header(ws, row, "Error Trend Over Time")
    chart3 = make_errors_trend_chart(bedrock_models)
    if chart3:
        row = insert_chart_image(ws, row, 1, chart3)
    else:
        ws.cell(row=row, column=1, value="No error trends to display.").font = SUBTITLE_FONT

    return ws


# =============================================================================
# Main
# =============================================================================
def main():
    if len(sys.argv) < 5:
        print("Usage: python3 generate_usage_charts.py <metrics_json> <output_dir> <start_date> <end_date>")
        sys.exit(1)

    metrics_json_path = sys.argv[1]
    output_dir = sys.argv[2]
    start_date = sys.argv[3]
    end_date = sys.argv[4]
    period_str = f"{start_date} to {end_date}"

    print("📊 Generating Bedrock Observability Excel Report...")
    print(f"   Period: {period_str}")

    # Load metrics data
    with open(metrics_json_path, 'r') as f:
        data = json.load(f)

    all_models = data.get('models', [])
    if not all_models:
        print("   ⚠️ No model data found in metrics file. Skipping report generation.")
        return

    # Separate bedrock runtime vs agents
    bedrock_models = [m for m in all_models if m.get('source') == 'bedrock']
    agent_models = [m for m in all_models if m.get('source') == 'agent']

    # Combine for charts (agents also have token/latency/error metrics)
    all_for_charts = bedrock_models + agent_models

    print(f"   Models: {len(bedrock_models)} Bedrock runtime + {len(agent_models)} Agent(s)")

    # Get AWS Account Info
    print("   Retrieving AWS account info...")
    account_id, account_name = get_aws_account_info()
    account_label = f"{account_name} ({account_id})" if account_name else account_id
    print(f"   Account: {account_label}")

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets
    print("   Building Sheet 1: Token Usage...")
    build_sheet_token_usage(wb, all_for_charts, account_label, period_str)

    print("   Building Sheet 2: Latency & Performance...")
    build_sheet_latency(wb, all_for_charts, account_label, period_str)

    print("   Building Sheet 3: Volume & Distribution...")
    build_sheet_volume(wb, all_for_charts, account_label, period_str)

    print("   Building Sheet 4: Reliability & Errors...")
    build_sheet_reliability(wb, all_for_charts, account_label, period_str)

    # Save Excel
    excel_path = os.path.join(output_dir, "Bedrock_Observability_Report.xlsx")
    wb.save(excel_path)
    print(f"   ✅ Excel saved: {excel_path}")

    # Create ZIP
    zip_path = os.path.join(output_dir, "bedrock_usage_report.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(excel_path, "Bedrock_Observability_Report.xlsx")
        # Also include the raw metrics JSON for reference
        zf.write(metrics_json_path, "raw_metrics/bedrock_all_metrics.json")

    print(f"   📦 ZIP created: {zip_path}")
    print("\n✅ Done!")


if __name__ == "__main__":
    main()
